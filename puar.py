import csv
import os
from argparse import ArgumentParser


def process_csv(input_file, template, prefix, postfix):
    with open(input_file) as ifo:
        csv_reader = csv.DictReader(ifo)
        if prefix:
            print(prefix)
        for row in csv_reader:
            print(template % row)
        if postfix:
            print(postfix)


def process_xls(input_file, template, prefix, postfix):
    try:
        import xlrd
    except ImportError:
        print('xlrd package is not installed, so XLS file processing is not supported')
        exit(1)
    workbook = xlrd.open_workbook(input_file)
    sheet = workbook.sheet_by_index(0)
    headers = sheet.row_values(0)
    if prefix:
        print(prefix)
    for row_num in range(1, sheet.nrows):
        row = dict(zip(headers, sheet.row_values(row_num)))
        print(template % row)
    if postfix:
        print(postfix)


def process_xlsx(input_file, template, prefix, postfix):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('openpyxl package is not installed, so XLSX file processing is not supported')
        exit(1)
    workbook = load_workbook(input_file)
    sheet = workbook.worksheets[0]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if prefix:
        print(prefix)
    for row_num in range(2, sheet.max_row + 1):
        row_values = [sheet.cell(row_num, column).value for column in range(1, sheet.max_column + 1)]
        row = dict(zip(headers, row_values))
        print(template % row)
    if postfix:
        print(postfix)


def main(input_file, template, prefix, postfix):
    ext = os.path.splitext(input_file)[-1].lower()
    if ext == '.csv':
        process_csv(input_file, template, prefix, postfix)
    elif ext == '.xls':
        process_xls(input_file, template, prefix, postfix)
    elif ext == '.xlsx':
        process_xlsx(input_file, template, prefix, postfix)
    else:
        print(f'Unknown filetype of "{args.input_file}"')
        exit(1)


if __name__ == '__main__':
    parser = ArgumentParser(description='Puar - Template based data transformation')
    parser.add_argument('input_file', help='Input file (csv, xls, xlsx)')
    parser.add_argument('template_string', help='printf-style format string')
    parser.add_argument('--prefix')
    parser.add_argument('--postfix')
    args = parser.parse_args()
    if os.path.isfile(args.input_file):
        main(args.input_file, args.template_string, args.prefix, args.postfix)
    else:
        print(f'Input file "{args.input_file}" not found')
        exit(1)
